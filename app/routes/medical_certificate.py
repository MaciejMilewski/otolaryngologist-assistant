import csv
import io
import logging
import os
import sqlite3
from datetime import datetime

import openpyxl
from bs4 import BeautifulSoup
from flask import Blueprint, render_template, abort, make_response, request, send_file, jsonify
from flask_login import current_user, login_required
from fpdf import FPDF
from jinja2 import TemplateNotFound

from app import parquet_data, dane_woj
from app.utils.const import typ_badan, place
from app.utils.medical_certificate_util import get_search_params, convert_row_to_dict, get_request_data, \
    pdf_orzeczenie_lekarskie, pdf_zaswiadczenie
from app.utils.parquet_util import get_streets_from_memory

medical_certificate_bp = Blueprint('medical_certificate', __name__)


@medical_certificate_bp.route('/medical_certificate', methods=['GET'])
@login_required
def medical_certificate_main():
    try:
        wojewodztwo_default = '22'  # Domyślne województwo - "POMORSKIE" (kod '22')

        if 'POMORSKIE' in parquet_data:
            miejscowosc_choices = parquet_data["POMORSKIE"]['Nazwa'].tolist()
        else:
            miejscowosc_choices = []

        default_miejscowosc = "Pruszcz Gdański"

        ulica_choices = get_streets_from_memory("POMORSKIE", default_miejscowosc)
        return render_template('medical_certificate.html',
                               user=current_user.login,
                               typ=typ_badan,
                               place=place,
                               woj=dane_woj,
                               woj_default=wojewodztwo_default,
                               cities=miejscowosc_choices,
                               city_default=default_miejscowosc,
                               streets=ulica_choices,
                               today=datetime.now().strftime('%Y-%m-%d'))
    except TemplateNotFound:
        return abort(404)


@medical_certificate_bp.route('/export_excel', methods=['POST'])
def export_excel():
    """
    Handles the export of search results to an Excel file.

    This endpoint processes the search parameters obtained through the helper
    function `get_search_params`, fetches the relevant data from the database
    using `fetch_data_from_db`, and generates an Excel file containing the results.
    The Excel file is then returned as a downloadable attachment.

    :return: A response containing the generated Excel file with the search results.
             The response sets the content disposition to 'attachment' with
             the filename 'results.xlsx' and the content type to
             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'.

    """
    # Wywołanie funkcji pomocniczej, która zwraca parametry wyszukiwania
    params = get_search_params()
    # Przekazanie parametrów do funkcji fetch_data_from_db
    # results = fetch_data_from_db(
    #     params['lastname'],
    #     params['start_date'],
    #     params['end_date'],
    #     params['zdolny'],
    #     params['all_type'],
    #     params['typ_badania']
    # )
    #
    # if results is None:
    #     return "Wystąpił błąd podczas przetwarzania zapytania.", 500
    #
    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = "Wyniki Badań"
    # ws.append(["Imię", "Nazwisko", "Data badania"])
    # for row in results:
    #     ws.append(row)
    # output = io.BytesIO()
    # wb.save(output)
    # output.seek(0)
    # response = make_response(output.getvalue())
    # response.headers["Content-Disposition"] = "attachment; filename=results.xlsx"
    # response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #
    # return response
    pass


@medical_certificate_bp.route('/export_csv', methods=['POST'])
def export_csv():
    """
    Handles the export of search results to a CSV file.

    :return: CSV file containing search results or an error message
    """
    # Wywołanie funkcji pomocniczej, która zwraca parametry wyszukiwania
    params = get_search_params()

    # Przekazanie parametrów do funkcji fetch_data_from_db


#     results = fetch_data_from_db(
#         params['lastname'],
#         params['start_date'],
#         params['end_date'],
#         params['zdolny'],
#         params['all_type'],
#         params['typ_badania']
#     )
#     if results is None:
#         logging.error("Export wyników wyszukiwania do pliku CSV nieudany - brak danych z bazy sqlite.")
#         return "Wystąpił błąd podczas przetwarzania zapytania.", 500
#
#     si = io.StringIO()
#     cw = csv.writer(si)
#     cw.writerow(["Imię", "Nazwisko", "Data badania"])
#     cw.writerows(results)
#     output = make_response(si.getvalue())
#     output.headers["Content-Disposition"] = "attachment; filename=results.csv"
#     output.headers["Content-type"] = "text/csv"
#
#     return output
#
#
@medical_certificate_bp.route('/generate_pdf')
def generate_pdf_for_raport():
    #     """
    #     Generates a PDF report based on the provided HTML table and selected test types within a specified date range.
    #     """
    #     # Pobierz dane z zapytania
    #     table_html = request.args.get('table_html', '')
    #     selected_types = request.args.getlist("selected_types")
    #     start_date = request.args.get("start_date")
    #     end_date = request.args.get("end_date")
    #
    #     # Słownik do mapowania typów badań na nazwy
    #     typ_badan_dict = {badanie['id']: badanie['name'] for badanie in typ_badan}
    #
    #     # Generowanie opisu wybranych typów badań
    #     if selected_types and selected_types != ['all']:
    #         badania_pdf = ", ".join(typ_badan_dict.get(typ, f"Typ {typ}") for typ in selected_types)
    #     else:
    #         badania_pdf = "wszystkie rodzaje badań !"
    #
    #     # Parsowanie HTML za pomocą BeautifulSoup
    #     soup = BeautifulSoup(table_html, 'html.parser')
    #
    #     # Znajdź indeksy kolumn "MIASTO" i "INFO"
    #     headers = [header.text.strip() for header in soup.find_all('th')]
    #     columns_to_remove = [headers.index("MIASTO"), headers.index("INFO")]
    #
    #     # Usuń nagłówki kolumn "MIASTO" i "INFO"
    #     headers = [header for i, header in enumerate(headers) if i not in columns_to_remove]
    #
    #     # Przetwórz wiersze i pomiń komórki odpowiadające "MIASTO" i "INFO"
    #     rows = [
    #         [cell.text.strip() for i, cell in enumerate(row.find_all('td')) if i not in columns_to_remove]
    #         for row in soup.find_all('tr')
    #     ]
    #
    #     # Tworzenie pliku PDF za pomocą FPDF
    #     pdf = FPDF()
    #     pdf.add_page()
    #
    #     # Dodanie czcionki TrueType
    #     pdf.add_font('DejaVu', '', './Fonts/DejaVuSans.ttf', uni=True)
    #     pdf.add_font('DejaVu-Bold', '', './Fonts/DejaVuSans-Bold.ttf', uni=True)
    #
    #     # Ustawienie czcionki podstawowej
    #     pdf.set_font("DejaVu", size=10)
    #
    #     # Dodanie logo i tytułu
    #     pdf.image('static/img/butterfly.png', 10, 8, 24, 24)
    #     pdf.set_font("DejaVu-Bold", size=18)
    #     pdf.cell(0, 10, "Raport", ln=True, align='C')  # Tytuł na środku
    #     pdf.ln(3)
    #
    #     # Dodanie informacji o zakresie dat i typach badań
    #     pdf.set_font("DejaVu", size=12)
    #     pdf.cell(0, 10, f"Zakres dni od {start_date} do {end_date}", ln=True, align='C')
    #     pdf.cell(0, 10, f"Typy badań: {badania_pdf}", ln=True, align='C')
    #
    #     # Dodanie odstępu przed tabelą
    #     pdf.ln(7)
    #
    #     # Tabela - nagłówki
    #     pdf.set_font("DejaVu-Bold", size=10)
    #     pdf.set_fill_color(60, 120, 180)  # Niebieski odcień dla nagłówków
    #     pdf.set_text_color(255, 255, 255)  # Biały tekst
    #     pdf.set_draw_color(50, 50, 50)  # Obramowanie ciemnoszare
    #     pdf.set_line_width(0.4)
    #
    #     # Usunięcie pustych wierszy z `rows`
    #     filtered_rows = [row for row in rows if row]  # Filtruje puste wiersze
    #
    #     # Pobierz szerokość strony PDF (domyślnie A4 w FPDF)
    #     page_width = pdf.w - 2 * pdf.l_margin  # Szerokość strony minus marginesy
    #
    #     # Obliczanie szerokości kolumn
    #     if filtered_rows:
    #         max_col_widths = [max(len(str(cell)) for cell in col) + 2 for col in zip(*([headers] + filtered_rows))]
    #     else:
    #         max_col_widths = [len(header) + 2 for header in headers]  # Gdy brak danych w wierszach
    #
    #     # Przeliczenie szerokości kolumn na procenty i dostosowanie do szerokości strony
    #     total_max_width = sum(max_col_widths)
    #     col_widths = [(page_width * w / total_max_width) for w in max_col_widths]
    #
    #     # Debugowanie
    #     # print("Page Width:", page_width)
    #     # print("Column Widths:", col_widths)
    #
    #     # Rysowanie nagłówków
    #     pdf.set_font("DejaVu-Bold", size=10)
    #     pdf.set_fill_color(60, 120, 180)  # Niebieskie tło
    #     pdf.set_text_color(255, 255, 255)  # Biały tekst
    #     pdf.set_draw_color(50, 50, 50)  # Obramowanie
    #     pdf.set_line_width(0.4)
    #
    #     for idx, header in enumerate(headers):
    #         pdf.cell(col_widths[idx], 10, header, border=1, align='C', fill=True)
    #     pdf.ln()
    #
    #     # Rysowanie wierszy
    #     pdf.set_font("DejaVu", size=10)  # Reset tekstu na czarny
    #     pdf.set_text_color(0, 0, 0)  # Czarny tekst
    #     for row_idx, row in enumerate(filtered_rows):
    #         fill_color = (245, 245, 245) if row_idx % 2 == 0 else (255, 255, 255)
    #         pdf.set_fill_color(*fill_color)
    #         for col_idx, cell in enumerate(row):
    #             pdf.cell(col_widths[col_idx], 10, cell, border=1, align='C', fill=True)
    #         pdf.ln()
    #
    #     # Zapisz plik PDF w pamięci
    #     pdf_data = pdf.output(dest='S')
    #     # Konwersja str na bytes za pomocą kodowania latin1
    #     pdf_bytes = pdf_data.encode('latin1')
    #
    #     buffer = io.BytesIO(pdf_bytes)
    #     buffer.seek(0)
    #
    #     # Zwróć plik PDF jako odpowiedź
    #     # Zwróć plik PDF jako odpowiedź
    #     return send_file(buffer, mimetype='application/pdf', download_name="raport.pdf", as_attachment=True)
    pass


@medical_certificate_bp.route('/get_miejscowosci', methods=['POST'])
def get_miejscowosci():
    """
    Fetches a list of localities (miejscowosci) for a given voivodeship code and returns them as pairs of (index, name).

    :return: A JSON object containing a list of localities.
    """
    kody_wojewodztwa = request.form.get('wojewodztwo', '')
    name_wojewodztwa = dane_woj.get(kody_wojewodztwa, '')

    # Pobierz tylko kolumnę 'Nazwa' z danych dla wybranego województwa
    if name_wojewodztwa in parquet_data:
        miejscowosci = parquet_data[name_wojewodztwa]['Nazwa'].tolist()  # Wyciągnięcie tylko kolumny 'Nazwa'
    else:
        miejscowosci = []

    # Zwróć listę miejscowości jako pary (index, nazwa)
    miejscowosci_pary = [(i, miejscowosc) for i, miejscowosc in enumerate(miejscowosci)]

    return jsonify({'miejscowosci': miejscowosci_pary})


@medical_certificate_bp.route('/get_person_details', methods=['POST'])
def get_person_details():
    """
    Handles POST requests to retrieve details of a person from the database.

    :return: JSON response containing person details if found,
             otherwise an error message with appropriate HTTP status code.
    """
    # try:
    #     person_id = get_request_data('id')
    #
    #     with sqlite3.connect(DATABASE) as connection:
    #         cursor = connection.cursor()
    #         cursor.execute("SELECT * FROM examined WHERE id = ?", (person_id,))
    #         row = cursor.fetchone()
    #     if row:
    #         person = convert_row_to_dict(row)
    #         return jsonify(person)
    #     else:
    #         logging.error({'error': 'Osoba nie została znaleziona'})
    #         return jsonify({'error': 'Osoba nie została znaleziona'}), 404
    # except ValueError as ve:
    #     logging.error({'error': str(ve)})
    #     return jsonify({'error': str(ve)}), 400
    # except sqlite3.DatabaseError as error_database:
    #     logging.error({'error': str(error_database)})
    #     return jsonify({'error': f'Błąd bazy danych: {str(error_database)}'}), 500
    # except Exception as error_global:
    #     logging.error({'error': str(error_global)})
    #     return jsonify({'error': f'Wystąpił błąd: {str(error_global)}'}), 500
    pass


@medical_certificate_bp.route('/get_ulice', methods=['POST'])
# Funkcja obsługująca AJAX do pobierania ulic
def get_ulice():
    """
    Handles the '/get_ulice' route to fetch street names via AJAX based on provided province and locality codes.

    :return: JSON response containing a list of street names and a status code. If the required province or locality codes are missing, it returns a 400 status. If no streets are found for the locality, it returns an empty list with a 200 status.
    """
    kod_wojewodztwa = request.form.get('wojewodztwo')
    if not kod_wojewodztwa:
        logging.error(f"/get_ulice : Brak kodów województwa aby szukać ulic!")
        return jsonify({'ulice': []}), 400

    kod_wojewodztwa = kod_wojewodztwa.strip()

    miejscowosc = request.form.get('miejscowosc')
    if not miejscowosc:
        logging.error(f"/get_ulice : Brak kodów miejscowości aby szukać ulic!")
        return jsonify({'ulice': []}), 400

    miejscowosc = miejscowosc.strip()

    name_wojewodztwa = dane_woj.get(kod_wojewodztwa, '')  # dict.get(key,'') zwraca wartość klucza albo ''

    result = get_streets_from_memory(name_wojewodztwa, miejscowosc)

    if result:
        return jsonify({'ulice': result}), 200  # Zwracamy wynik w formacie JSON i status 200
    else:
        logging.info(f"/get_ulice : Brak ulic dla miejscowości o symbolu {miejscowosc}.")
        return jsonify({'ulice': []}), 200  # Zwracamy pustą listę ulic, ale status OK


@medical_certificate_bp.route('/pdf', methods=['POST'])
def generate_pdf():
    """Generuje plik PDF na podstawie danych przesłanych w żądaniu POST."""
    try:
        data = request.json

        # Tworzenie katalogu na pliki PDF
        base_dir = os.path.dirname(os.path.abspath(__file__))
        STATIC_PDF_DIR = os.path.join(base_dir, 'pdf')
        pdf_dir = STATIC_PDF_DIR
        os.makedirs(pdf_dir, exist_ok=True)

        # Nazwa pliku PDF
        file_pdf_name = f"{data['lastname']}-{data['firstname']}.pdf"
        pdf_path = os.path.join(pdf_dir, file_pdf_name)

        if data['typ'] == '11':
            pdf = pdf_orzeczenie_lekarskie(data)
        else:
            pdf = pdf_zaswiadczenie(data)

        # Zapis PDF
        pdf.output(pdf_path)

        # Tworzenie URL do pliku PDF
        pdf_url = pdf_path.replace("\\", "/")

        return jsonify({'file': pdf_url})
    except Exception as error_building:
        logging.error(f"Błąd podczas generowania PDF: {str(error_building)}")
        return jsonify({'błąd': 'Nie udało się wygenerować PDF'}), 500
